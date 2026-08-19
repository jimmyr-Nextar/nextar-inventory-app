import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from io import BytesIO
from datetime import date
 
st.set_page_config(page_title="Nextar Inventory Report", page_icon="📦", layout="centered")
 
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.main { background: #F8FAFC; }
.report-header { background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%); color: white; padding: 2rem 2.5rem; border-radius: 12px; margin-bottom: 2rem; }
.report-header h1 { font-size: 1.8rem; font-weight: 700; margin: 0; letter-spacing: -0.5px; }
.report-header p { margin: 0.3rem 0 0; opacity: 0.8; font-size: 0.95rem; font-weight: 300; }
.card { background: white; border-radius: 10px; padding: 1.5rem; border: 1px solid #E2E8F0; margin-bottom: 1rem; }
.card h3 { font-size: 0.85rem; font-weight: 600; color: #64748B; text-transform: uppercase; letter-spacing: 0.08em; margin: 0 0 1rem; }
.stButton > button { background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%) !important; color: white !important; border: none !important; border-radius: 8px !important; padding: 0.6rem 2rem !important; font-weight: 600 !important; font-size: 0.95rem !important; width: 100% !important; }
.warning-box { background: #FFF8E1; border-left: 4px solid #F59E0B; padding: 0.8rem 1rem; border-radius: 0 8px 8px 0; font-size: 0.875rem; color: #78350F; margin-bottom: 1rem; }
.info-box { background: #EFF6FF; border-left: 4px solid #2E75B6; padding: 0.8rem 1rem; border-radius: 0 8px 8px 0; font-size: 0.875rem; color: #1E3A5F; margin-bottom: 1rem; }
.success-box { background: #F0FDF4; border-left: 4px solid #22C55E; padding: 0.8rem 1rem; border-radius: 0 8px 8px 0; font-size: 0.875rem; color: #14532D; margin-bottom: 1rem; }
.transfer-box { background: #F0FDF4; border-left: 4px solid #16A34A; padding: 0.8rem 1rem; border-radius: 0 8px 8px 0; font-size: 0.875rem; color: #14532D; margin-bottom: 1rem; }
</style>
""", unsafe_allow_html=True)
 
EOL_ITEMS = ["APL IPHONE 13 128G BLK TMUS KIT", "TCL T440W FLIP 4 5G 32G GRY KIT"]
 
def build_report(df, market_name, days, report_date, transfers_enabled, transfer_excluded):
    MIN_DAYS = 10
    OVERSTOCK_DAYS = 30
    stores = sorted(df["custno"].unique())
    palette = ["EBF3FB","FFF2CC","E2EFDA","FCE4D6","E8DAEF","D5F5E3","FDEBD0","D6EAF8","FDEDEC","E8F8F5","F9EBEA","EAF2FF","FEF9E7","F4ECF7","F0F3F4","FAF5FF","FFF9F0","F0FFF4"]
    store_colors = {s: palette[i % len(palette)] for i, s in enumerate(stores)}
 
    df = df.copy()
    df["daily_rate"] = df["totsold"] / days
    df["days_on_hand"] = df.apply(lambda r: round(r["onhand"] / r["daily_rate"], 1) if r["daily_rate"] > 0 else 999, axis=1)
    df["target_qty"] = (df["daily_rate"] * MIN_DAYS).round(1)
    df["order_qty"] = df.apply(lambda r: max(1, round(r["target_qty"] - r["onhand"])) if r["onhand"] == 0 else max(0, round(r["target_qty"] - r["onhand"])), axis=1)
    df["surplus"] = df.apply(lambda r: max(0, r["onhand"] - round(r["target_qty"])), axis=1)
    df["is_eol"] = df["itmdesc"].isin(EOL_ITEMS)
    df["status"] = df.apply(lambda r: "EOL - TRACK ONLY" if r["is_eol"] and r["order_qty"] > 0 else ("ORDER NOW" if r["order_qty"] > 0 else ("WATCH" if r["days_on_hand"] < 14 else "OK")), axis=1)
 
    summary = df.groupby(["item","itmdesc"]).agg(total_onhand=("onhand","sum"), total_sold=("totsold","sum"), total_order_qty=("order_qty","sum")).reset_index()
    summary["daily_rate"] = (summary["total_sold"] / days).round(2)
    summary["days_on_hand"] = summary.apply(lambda r: round(r["total_onhand"] / r["daily_rate"], 1) if r["daily_rate"] > 0 else 999, axis=1)
    summary["target_qty"] = (summary["daily_rate"] * MIN_DAYS).round(1)
    summary["order_qty"] = summary["total_order_qty"]
    summary["is_eol"] = summary["itmdesc"].isin(EOL_ITEMS)
    summary["status"] = summary.apply(lambda r: "EOL - TRACK ONLY" if r["is_eol"] and r["order_qty"] > 0 else ("ORDER NOW" if r["order_qty"] > 0 else ("WATCH" if r["days_on_hand"] < 14 else "OK")), axis=1)
    summary = summary.sort_values(["order_qty","days_on_hand"], ascending=[False,True])
 
    store_detail = df[["custno","company","itmdesc","item","onhand","totsold","daily_rate","days_on_hand","target_qty","order_qty","surplus","status","is_eol"]].copy()
    store_detail = store_detail.sort_values(["custno","order_qty","days_on_hand"], ascending=[True,False,True])
    order_rows = store_detail[(store_detail["order_qty"] > 0) & (~store_detail["is_eol"])].copy()
    eol_rows = store_detail[(store_detail["is_eol"]) & (store_detail["order_qty"] > 0)].copy()
    total_order = int(order_rows["order_qty"].sum())
 
    transfers = []
    has_surplus = df[(df["surplus"] > 0) & (df["days_on_hand"] > OVERSTOCK_DAYS) & (~df["custno"].isin(transfer_excluded))].copy()
    if transfers_enabled:
        needs = df[(df["order_qty"] > 0) & (~df["custno"].isin(transfer_excluded)) & (~df["is_eol"])].copy()
        for _, need_row in needs.iterrows():
            remaining_need = need_row["order_qty"]
            for _, surp_row in has_surplus.iterrows():
                if need_row["item"] == surp_row["item"] and need_row["custno"] != surp_row["custno"] and remaining_need > 0:
                    avail = surp_row["surplus"]
                    if avail > 0:
                        transfer_qty = min(remaining_need, avail)
                        transfers.append({"from_store": surp_row["custno"], "from_address": surp_row["company"], "to_store": need_row["custno"], "to_address": need_row["company"], "item": str(need_row["item"]).replace("'",""), "description": need_row["itmdesc"], "transfer_qty": int(transfer_qty), "remaining_need_after": int(remaining_need - transfer_qty), "from_days_on_hand": surp_row["days_on_hand"], "to_days_on_hand": need_row["days_on_hand"]})
                        remaining_need -= transfer_qty
 
    RED = "FFC7CE"; RED_F = "9C0006"; YELLOW = "FFEB9C"; YELLOW_F = "9C6500"
    GREEN = "C6EFCE"; GREEN_F = "276221"; EOL_BG = "F2DCDB"; EOL_F = "833C00"
    HDR_BG = "1F4E79"; HDR_F = "FFFFFF"; BANNER = "1F4E79"
    thin = Side(style="thin", color="BFBFBF")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    date_str = report_date.strftime("%B %d, %Y")
 
    def mh(ws, row, c1, c2):
        for c in range(c1, c2+1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(bold=True, color=HDR_F, name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=HDR_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
 
    def sc(s):
        if s == "ORDER NOW": return RED, RED_F
        if s == "WATCH": return YELLOW, YELLOW_F
        if s == "EOL - TRACK ONLY": return EOL_BG, EOL_F
        return GREEN, GREEN_F
 
    if transfers_enabled:
        xfer_note = f"Transfers ON | Excluded: {', '.join(transfer_excluded)}" if transfer_excluded else "Transfers ON - all stores eligible"
        over_note = f"Excluded from transfers: {', '.join(transfer_excluded)}" if transfer_excluded else "All stores eligible for transfers."
        over_color = "FCE4D6" if transfer_excluded else "D6E4F0"
        over_fc = "833C00" if transfer_excluded else "1F4E79"
    else:
        xfer_note = "Transfers not enabled"
        over_note = "Transfers are not enabled for this report."
        over_color = "D6E4F0"; over_fc = "1F4E79"
 
    wb = Workbook()
 
    # Sheet 1: Order Summary
    ws1 = wb.active; ws1.title = "Order Summary"
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"INVENTORY REORDER REPORT  -  {market_name.upper()}  -  {date_str}  |  10-Day Minimum"
    ws1["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws1["A1"].fill = PatternFill("solid", start_color=BANNER)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28
    ws1.merge_cells("A2:I2")
    ws1["A2"] = f"Report Date: {date_str}   |   {days}-Day Analysis   |   Min Days on Hand: 10   |   Stores: {len(stores)}   |   SKUs: {df['item'].nunique()}   |   {xfer_note}"
    ws1["A2"].font = Font(italic=True, name="Arial", size=9, color="404040")
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1["A2"].fill = PatternFill("solid", start_color="D6E4F0")
    ws1.row_dimensions[2].height = 18
    ws1.append([])
    ws1.append(["Item #","Description",f"On Hand\n(All Stores)",f"{days}-Day\nSales","Daily\nSell Rate","Days\nOn Hand","10-Day\nTarget","Order Qty\n(All Stores)","Status"])
    mh(ws1, 4, 1, 9); ws1.row_dimensions[4].height = 36
 
    for i, (_, row) in enumerate(summary.iterrows()):
        r = ws1.max_row + 1
        dd = str(row["days_on_hand"]) if row["days_on_hand"] < 999 else "-"
        ws1.append([str(row["item"]).replace("'",""), row["itmdesc"], row["total_onhand"], row["total_sold"], row["daily_rate"], dd, row["target_qty"], row["order_qty"], row["status"]])
        bg, fg = sc(row["status"])
        alt = "F7FBFF" if i % 2 else "FFFFFF"
        for col in range(1, 10):
            cell = ws1.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10, color="595959" if row["is_eol"] else "000000", italic=bool(row["is_eol"]))
            cell.fill = PatternFill("solid", start_color=alt)
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
            cell.border = bdr
        if row["order_qty"] > 0:
            c8 = ws1.cell(row=r, column=8)
            c8.font = Font(bold=True, name="Arial", size=10, color=EOL_F if row["is_eol"] else RED_F)
            c8.fill = PatternFill("solid", start_color=EOL_BG if row["is_eol"] else RED)
        c9 = ws1.cell(row=r, column=9)
        c9.fill = PatternFill("solid", start_color=bg); c9.font = Font(bold=True, name="Arial", size=10, color=fg)
        c9.alignment = Alignment(horizontal="center", vertical="center"); c9.border = bdr
        ws1.row_dimensions[r].height = 18
 
    for i, w in enumerate([18,40,13,12,12,12,12,13,15], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A5"
    lr = ws1.max_row + 2
    ws1.cell(row=lr, column=1).value = "STATUS KEY:"; ws1.cell(row=lr, column=1).font = Font(bold=True, name="Arial", size=9)
    for col, (lbl, bg, fg) in enumerate([("ORDER NOW - Below 10-day target or 0 on hand", RED, RED_F), ("WATCH - Under 14 days on hand", YELLOW, YELLOW_F), ("EOL - TRACK ONLY - Do not purchase", EOL_BG, EOL_F), ("OK - Adequate stock", GREEN, GREEN_F)], 2):
        c = ws1.cell(row=lr, column=col); c.value = lbl
        c.fill = PatternFill("solid", start_color=bg); c.font = Font(bold=True, name="Arial", size=9, color=fg)
        c.alignment = Alignment(horizontal="center"); c.border = bdr
 
    # Sheet 2: By Store
    ws2 = wb.create_sheet("By Store")
    ws2.merge_cells("A1:K1")
    ws2["A1"] = f"STORE-LEVEL ORDER DETAIL  -  {market_name.upper()}  -  {date_str}"
    ws2["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws2["A1"].fill = PatternFill("solid", start_color=BANNER)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28; ws2.append([])
    ws2.append(["Store ID","Store Name","Item #","Description","On Hand",f"{days}-Day\nSales","Daily\nRate","Days\nOn Hand","10-Day\nTarget","Order\nQty","Status"])
    mh(ws2, 3, 1, 11); ws2.row_dimensions[3].height = 36
 
    for i, (_, row) in enumerate(store_detail.iterrows()):
        r = ws2.max_row + 1
        bg2 = store_colors.get(row["custno"], "FFFFFF")
        dd = str(row["days_on_hand"]) if row["days_on_hand"] < 999 else "-"
        ws2.append([row["custno"], row["company"], str(row["item"]).replace("'",""), row["itmdesc"], row["onhand"], row["totsold"], round(row["daily_rate"],2), dd, row["target_qty"], int(row["order_qty"]), row["status"]])
        sbg, sfg = sc(row["status"])
        for col in range(1, 12):
            cell = ws2.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10, italic=bool(row["is_eol"]))
            cell.fill = PatternFill("solid", start_color=bg2)
            cell.alignment = Alignment(horizontal="center" if col not in [2,4] else "left", vertical="center"); cell.border = bdr
        if row["order_qty"] > 0:
            c10 = ws2.cell(row=r, column=10)
            c10.font = Font(bold=True, name="Arial", size=10, color=EOL_F if row["is_eol"] else RED_F)
            c10.fill = PatternFill("solid", start_color=EOL_BG if row["is_eol"] else RED)
        c11 = ws2.cell(row=r, column=11)
        c11.fill = PatternFill("solid", start_color=sbg); c11.font = Font(bold=True, name="Arial", size=10, color=sfg)
        c11.alignment = Alignment(horizontal="center", vertical="center"); c11.border = bdr
        ws2.row_dimensions[r].height = 18
 
    for i, w in enumerate([13,28,18,38,10,11,10,12,12,11,16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"
 
    # Sheet 3: Order List
    ws3 = wb.create_sheet("Order List")
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "ITEMS TO ORDER  -  Below 10-day minimum OR 0 on hand  |  EOL items excluded"
    ws3["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws3["A1"].fill = PatternFill("solid", start_color="C00000")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28; ws3.append([])
    ws3.append(["Store ID","Store Name","Item #","Description","On Hand","Days on Hand","Order Qty"])
    mh(ws3, 3, 1, 7); ws3.row_dimensions[3].height = 30
 
    for i, (_, row) in enumerate(order_rows.iterrows()):
        r = ws3.max_row + 1; alt = i % 2 == 1
        dd = str(row["days_on_hand"]) if row["days_on_hand"] < 999 else "-"
        ws3.append([row["custno"], row["company"], str(row["item"]).replace("'",""), row["itmdesc"], row["onhand"], dd, int(row["order_qty"])])
        for col in range(1, 8):
            cell = ws3.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=RED if not alt else "FFD7D7")
            cell.alignment = Alignment(horizontal="center" if col not in [2,4] else "left", vertical="center"); cell.border = bdr
        ws3.cell(row=r, column=7).font = Font(bold=True, name="Arial", size=10, color=RED_F)
        if row["onhand"] == 0:
            ws3.cell(row=r, column=5).font = Font(bold=True, name="Arial", size=10, color=RED_F)
        ws3.row_dimensions[r].height = 18
 
    for i, w in enumerate([13,28,18,40,10,14,12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A4"
    tr = ws3.max_row + 2
    ws3.cell(row=tr, column=6).value = "TOTAL UNITS TO ORDER:"; ws3.cell(row=tr, column=6).font = Font(bold=True, name="Arial", size=11)
    ws3.cell(row=tr, column=6).alignment = Alignment(horizontal="right")
    ws3.cell(row=tr, column=7).value = total_order; ws3.cell(row=tr, column=7).font = Font(bold=True, name="Arial", size=11, color=RED_F)
    ws3.cell(row=tr, column=7).fill = PatternFill("solid", start_color=RED); ws3.cell(row=tr, column=7).border = bdr
    ws3.cell(row=tr, column=7).alignment = Alignment(horizontal="center")
 
    # Sheet 4: EOL
    ws_eol = wb.create_sheet("EOL Watch List")
    ws_eol.merge_cells("A1:G1")
    ws_eol["A1"] = "EOL WATCH LIST  -  Do NOT purchase. Order only if stock becomes available."
    ws_eol["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws_eol["A1"].fill = PatternFill("solid", start_color="833C00")
    ws_eol["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_eol.row_dimensions[1].height = 28
    ws_eol.merge_cells("A2:G2")
    ws_eol["A2"] = "EOL: APL iPhone 13 128G BLK   |   TCL T440W Flip 4 5G (replaced by TCL Flip 4 V2 - SKU 610214726322)"
    ws_eol["A2"].font = Font(italic=True, name="Arial", size=9, color="595959")
    ws_eol["A2"].fill = PatternFill("solid", start_color="FCE4D6")
    ws_eol["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_eol.row_dimensions[2].height = 18; ws_eol.append([])
    ws_eol.append(["Store ID","Store Name","Item #","Description","On Hand","Days on Hand","Would Order\n(if available)"])
    mh(ws_eol, 4, 1, 7); ws_eol.row_dimensions[4].height = 36
 
    if eol_rows.empty:
        ws_eol.merge_cells("A5:G5")
        ws_eol["A5"] = "No EOL items currently below target - all EOL stock is adequate."
        ws_eol["A5"].font = Font(italic=True, name="Arial", size=10, color="276221")
        ws_eol["A5"].fill = PatternFill("solid", start_color=GREEN)
        ws_eol["A5"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        for i, (_, row) in enumerate(eol_rows.iterrows()):
            r = ws_eol.max_row + 1; alt = i % 2 == 1
            dd = str(row["days_on_hand"]) if row["days_on_hand"] < 999 else "-"
            ws_eol.append([row["custno"], row["company"], str(row["item"]).replace("'",""), row["itmdesc"], row["onhand"], dd, int(row["order_qty"])])
            for col in range(1, 8):
                cell = ws_eol.cell(row=r, column=col)
                cell.font = Font(name="Arial", size=10, italic=True, color="595959")
                cell.fill = PatternFill("solid", start_color=EOL_BG if not alt else "FAE9E8")
                cell.alignment = Alignment(horizontal="center" if col not in [2,4] else "left", vertical="center"); cell.border = bdr
            ws_eol.cell(row=r, column=7).font = Font(bold=True, name="Arial", size=10, color=EOL_F)
            ws_eol.row_dimensions[r].height = 18
        te = ws_eol.max_row + 2
        ws_eol.cell(row=te, column=6).value = "TOTAL IF AVAILABLE:"; ws_eol.cell(row=te, column=6).font = Font(bold=True, name="Arial", size=11)
        ws_eol.cell(row=te, column=6).alignment = Alignment(horizontal="right")
        ws_eol.cell(row=te, column=7).value = int(eol_rows["order_qty"].sum())
        ws_eol.cell(row=te, column=7).font = Font(bold=True, name="Arial", size=11, color=EOL_F)
        ws_eol.cell(row=te, column=7).fill = PatternFill("solid", start_color=EOL_BG); ws_eol.cell(row=te, column=7).border = bdr
        ws_eol.cell(row=te, column=7).alignment = Alignment(horizontal="center")
    for i, w in enumerate([13,28,18,40,10,14,16], 1):
        ws_eol.column_dimensions[get_column_letter(i)].width = w
    ws_eol.freeze_panes = "A5"
 
    # Sheet 5: Overstock
    ws_over = wb.create_sheet("Overstock Inventory")
    ws_over.merge_cells("A1:H1")
    ws_over["A1"] = "OVERSTOCK INVENTORY  -  Items exceeding 30-day supply"
    ws_over["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws_over["A1"].fill = PatternFill("solid", start_color="7030A0")
    ws_over["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_over.row_dimensions[1].height = 28
    ws_over.merge_cells("A2:H2")
    ws_over["A2"] = over_note; ws_over["A2"].font = Font(italic=True, name="Arial", size=9, color=over_fc)
    ws_over["A2"].fill = PatternFill("solid", start_color=over_color)
    ws_over["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_over.row_dimensions[2].height = 18; ws_over.append([])
    ws_over.append(["Store ID","Store Name","Item #","Description","On Hand","Target\nQty (10d)","Surplus\nUnits","Days\nOn Hand"])
    mh(ws_over, 4, 1, 8); ws_over.row_dimensions[4].height = 36
 
    over_rows = store_detail[(store_detail["surplus"] > 0) & (store_detail["days_on_hand"] > OVERSTOCK_DAYS)].copy()
    over_rows = over_rows.sort_values(["custno","days_on_hand"], ascending=[True,False])
    for i, (_, row) in enumerate(over_rows.iterrows()):
        r = ws_over.max_row + 1
        bg2 = store_colors.get(row["custno"], "EBF3FB")
        dd = "-" if row["days_on_hand"] >= 999 else str(row["days_on_hand"])
        ws_over.append([row["custno"], row["company"], str(row["item"]).replace("'",""), row["itmdesc"], row["onhand"], row["target_qty"], int(row["surplus"]), dd])
        for col in range(1, 9):
            cell = ws_over.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10); cell.fill = PatternFill("solid", start_color=bg2)
            cell.alignment = Alignment(horizontal="center" if col not in [2,4] else "left", vertical="center"); cell.border = bdr
        ws_over.cell(row=r, column=7).font = Font(bold=True, name="Arial", size=10, color="7030A0")
        ws_over.row_dimensions[r].height = 18
    for i, w in enumerate([13,28,18,38,10,13,12,13], 1):
        ws_over.column_dimensions[get_column_letter(i)].width = w
    ws_over.freeze_panes = "A4"
 
    # Sheet 6: Transfer Plan
    ws_xfer = wb.create_sheet("Transfer Plan")
    ws_xfer.merge_cells("A1:I1")
    ws_xfer["A1"] = "INVENTORY TRANSFER PLAN  -  Move overstock to cover shortfalls"
    ws_xfer["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws_xfer["A1"].fill = PatternFill("solid", start_color="375623")
    ws_xfer["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_xfer.row_dimensions[1].height = 28
    ws_xfer.merge_cells("A2:I2")
    ws_xfer["A2"] = over_note; ws_xfer["A2"].font = Font(italic=True, name="Arial", size=9, color=over_fc)
    ws_xfer["A2"].fill = PatternFill("solid", start_color=over_color)
    ws_xfer["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_xfer.row_dimensions[2].height = 18; ws_xfer.append([])
 
    if not transfers_enabled:
        ws_xfer.merge_cells("A4:I4"); ws_xfer["A4"] = "Transfers are not enabled for this report."
        ws_xfer["A4"].font = Font(italic=True, name="Arial", size=10); ws_xfer["A4"].alignment = Alignment(horizontal="center")
    elif not transfers:
        ws_xfer.merge_cells("A4:I4"); ws_xfer["A4"] = "No inter-store transfers possible at this time."
        ws_xfer["A4"].font = Font(italic=True, name="Arial", size=10); ws_xfer["A4"].alignment = Alignment(horizontal="center")
    else:
        ws_xfer.append(["Item #","Description","Transfer\nQty","FROM Store","FROM Address","FROM Days\nOn Hand","TO Store","TO Address","Remaining\nOrder Need"])
        mh(ws_xfer, 4, 1, 9); ws_xfer.row_dimensions[4].height = 36
        for i, t in enumerate(transfers):
            r = ws_xfer.max_row + 1; bg2 = "E2EFDA" if i % 2 == 0 else "F2F9EE"
            ws_xfer.append([t["item"], t["description"], t["transfer_qty"], t["from_store"], t["from_address"], t["from_days_on_hand"], t["to_store"], t["to_address"], t["remaining_need_after"]])
            for col in range(1, 10):
                cell = ws_xfer.cell(row=r, column=col)
                cell.font = Font(name="Arial", size=10); cell.fill = PatternFill("solid", start_color=bg2)
                cell.alignment = Alignment(horizontal="center" if col not in [2,5,8] else "left", vertical="center"); cell.border = bdr
            tq = ws_xfer.cell(row=r, column=3)
            tq.font = Font(bold=True, name="Arial", size=11, color=GREEN_F); tq.fill = PatternFill("solid", start_color=GREEN)
            rn = ws_xfer.cell(row=r, column=9)
            if t["remaining_need_after"] > 0:
                rn.font = Font(bold=True, name="Arial", size=10, color=RED_F); rn.fill = PatternFill("solid", start_color=RED)
            else:
                rn.font = Font(bold=True, name="Arial", size=10, color=GREEN_F); rn.fill = PatternFill("solid", start_color=GREEN)
            ws_xfer.row_dimensions[r].height = 18
        for i, w in enumerate([18,36,12,13,26,15,13,26,16], 1):
            ws_xfer.column_dimensions[get_column_letter(i)].width = w
        ws_xfer.freeze_panes = "A5"
        total_xfer = sum(t["transfer_qty"] for t in transfers)
        remaining_orders = sum(t["remaining_need_after"] for t in transfers)
        sr = ws_xfer.max_row + 2
        ws_xfer.merge_cells(f"A{sr}:C{sr}")
        ws_xfer.cell(row=sr, column=1).value = "TRANSFER SAVINGS SUMMARY"
        ws_xfer.cell(row=sr, column=1).font = Font(bold=True, name="Arial", size=10, color=GREEN_F)
        ws_xfer.cell(row=sr, column=1).fill = PatternFill("solid", start_color=GREEN)
        ws_xfer.cell(row=sr, column=1).alignment = Alignment(horizontal="center"); ws_xfer.cell(row=sr, column=1).border = bdr
        for offset, (label, val, use_red) in enumerate([("Units moved via transfer (no new order needed):", total_xfer, False), ("Units still needing external purchase:", remaining_orders, remaining_orders > 0)], 1):
            rn2 = sr + offset
            ws_xfer.cell(row=rn2, column=1).value = label; ws_xfer.cell(row=rn2, column=1).font = Font(name="Arial", size=10)
            ws_xfer.cell(row=rn2, column=3).value = val
            fc = RED_F if use_red else GREEN_F; bg2 = RED if use_red else GREEN
            ws_xfer.cell(row=rn2, column=3).font = Font(bold=True, name="Arial", size=11, color=fc)
            ws_xfer.cell(row=rn2, column=3).fill = PatternFill("solid", start_color=bg2)
            ws_xfer.cell(row=rn2, column=3).alignment = Alignment(horizontal="center"); ws_xfer.cell(row=rn2, column=3).border = bdr
 
    # Sheet 7: SKU Reference
    ws_sku = wb.create_sheet("SKU Reference")
    ws_sku.merge_cells("A1:D1")
    ws_sku["A1"] = f"SKU REFERENCE  -  {market_name.upper()}  -  All Active SKUs"
    ws_sku["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws_sku["A1"].fill = PatternFill("solid", start_color="404040")
    ws_sku["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sku.row_dimensions[1].height = 28; ws_sku.append([])
    ws_sku.append(["SKU / Item #","System Description",f"Total On Hand\n(All Stores)",f"{days}-Day\nTotal Sales"])
    mh(ws_sku, 3, 1, 4); ws_sku.row_dimensions[3].height = 36
 
    sku_ref = df.groupby(["item","itmdesc"]).agg(total_onhand=("onhand","sum"), total_sold=("totsold","sum")).reset_index().sort_values("itmdesc")
    for i, (_, row) in enumerate(sku_ref.iterrows()):
        r = ws_sku.max_row + 1; is_eol = row["itmdesc"] in EOL_ITEMS
        ws_sku.append([str(row["item"]).replace("'",""), row["itmdesc"], row["total_onhand"], row["total_sold"]])
        for col in range(1, 5):
            cell = ws_sku.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10, italic=is_eol, color="595959" if is_eol else "000000")
            cell.fill = PatternFill("solid", start_color=EOL_BG if is_eol else ("F7FBFF" if i%2 else "FFFFFF"))
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center"); cell.border = bdr
        ws_sku.row_dimensions[r].height = 18
    for i, w in enumerate([20,44,15,14], 1):
        ws_sku.column_dimensions[get_column_letter(i)].width = w
    ws_sku.freeze_panes = "A4"
    nr = ws_sku.max_row + 2; ws_sku.merge_cells(f"A{nr}:D{nr}")
    ws_sku.cell(row=nr, column=1).value = "EOL items italicized. TCL T440W Flip 4 5G EOL - replaced by TCL Flip 4 V2 32G GRY (SKU 610214726322)."
    ws_sku.cell(row=nr, column=1).font = Font(italic=True, name="Arial", size=9, color="595959")
    ws_sku.cell(row=nr, column=1).alignment = Alignment(horizontal="center")
 
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf, {"order_lines": len(order_rows), "total_units": total_order, "transfers": len(transfers), "overstock": len(over_rows), "eol_lines": len(eol_rows), "stores": len(stores), "transfers_data": transfers, "total_xfer": sum(t["transfer_qty"] for t in transfers) if transfers else 0, "remaining_orders": sum(t["remaining_need_after"] for t in transfers) if transfers else 0}
 
 
# UI
st.markdown('<div class="report-header"><h1>📦 Nextar Inventory Report</h1><p>Upload your sales file, configure your settings, and download your formatted order report.</p></div>', unsafe_allow_html=True)
 
st.markdown('<div class="card"><h3>Step 1 — Market Name</h3>', unsafe_allow_html=True)
market_name = st.text_input("Market name", placeholder="e.g. Los Angeles, New Mexico, Reno, Tucson", label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)
 
st.markdown('<div class="card"><h3>Step 2 — Sales Window</h3>', unsafe_allow_html=True)
days = st.select_slider("Sales window", options=[7, 14, 21], value=14, label_visibility="collapsed")
st.markdown(f'<div class="info-box">Using <strong>{days}-day</strong> sales window to calculate daily run rate.</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
 
st.markdown('<div class="card"><h3>Step 3 — Transfer Settings</h3>', unsafe_allow_html=True)
transfers_enabled = st.toggle("Enable inter-store transfer recommendations", value=False)
transfer_excluded = []
if transfers_enabled:
    st.markdown('<div class="transfer-box">Transfer plan will be included in your report.</div>', unsafe_allow_html=True)
    excluded_input = st.text_input("Stores to exclude from transfers (optional — leave blank if all stores are eligible)", placeholder="e.g. NTG-NM03, NTG-NM04, NTG-LA96")
    if excluded_input.strip():
        transfer_excluded = [s.strip() for s in excluded_input.split(",") if s.strip()]
        st.markdown(f'<div class="warning-box">Excluded from transfers: <strong>{", ".join(transfer_excluded)}</strong></div>', unsafe_allow_html=True)
else:
    st.markdown('<div class="info-box">Transfers off. Toggle on to include transfer recommendations in your report.</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
 
st.markdown('<div class="card"><h3>Step 4 — Upload Sales File</h3>', unsafe_allow_html=True)
uploaded = st.file_uploader("Upload sales Excel file (.xlsx)", type=["xlsx"], label_visibility="collapsed")
df = None
if uploaded:
    try:
        df = pd.read_excel(uploaded)
        required = ["custno","company","item","itmdesc","onhand","totsold"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            st.markdown(f'<div class="warning-box">Missing columns: {", ".join(missing)}. Please check your file.</div>', unsafe_allow_html=True)
            df = None
        else:
            st.markdown(f'<div class="success-box">File loaded — <strong>{df["custno"].nunique()} stores</strong>, <strong>{df["item"].nunique()} SKUs</strong>, <strong>{len(df)} rows</strong></div>', unsafe_allow_html=True)
    except Exception as e:
        st.markdown(f'<div class="warning-box">Could not read file: {str(e)}</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
 
st.markdown('<div class="card"><h3>Step 5 — Report Date</h3>', unsafe_allow_html=True)
report_date = st.date_input("Report date", value=date.today(), label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)
 
st.markdown('<div class="card"><h3>Step 6 — Generate Report</h3>', unsafe_allow_html=True)
ready = df is not None and bool(market_name.strip())
if ready:
    xfer_label = "Transfers ON" if transfers_enabled else "Transfers OFF"
    if transfers_enabled and transfer_excluded:
        xfer_label += f" | Excluding: {', '.join(transfer_excluded)}"
    st.markdown(f'<div class="info-box">Ready — <strong>{market_name}</strong> | <strong>{days}-day</strong> window | <strong>10-day</strong> min | {xfer_label}</div>', unsafe_allow_html=True)
else:
    items = []
    if not market_name.strip(): items.append("market name")
    if df is None: items.append("sales file")
    st.markdown(f'<div class="info-box">Please enter a {" and ".join(items)} above to continue.</div>', unsafe_allow_html=True)
 
if st.button("Run Report", disabled=not ready, use_container_width=True):
    with st.spinner("Building your report..."):
        try:
            buf, stats = build_report(df, market_name.strip(), days, report_date, transfers_enabled, transfer_excluded)
            st.markdown("---")
            st.markdown("### Report Summary")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.markdown(f'<div class="metric-box"><div class="num red">{stats["order_lines"]}</div><div class="lbl">Order Lines</div></div>', unsafe_allow_html=True)
            with col2:
                st.markdown(f'<div class="metric-box"><div class="num red">{stats["total_units"]}</div><div class="lbl">Units to Order</div></div>', unsafe_allow_html=True)
            with col3:
                st.markdown(f'<div class="metric-box"><div class="num">{stats["transfers"]}</div><div class="lbl">Transfers Found</div></div>', unsafe_allow_html=True)
            with col4:
                st.markdown(f'<div class="metric-box"><div class="num">{stats["overstock"]}</div><div class="lbl">Overstock SKUs</div></div>', unsafe_allow_html=True)
            if stats["eol_lines"] > 0:
                st.markdown(f'<div class="warning-box">EOL Watch List — {stats["eol_lines"]} item(s) below target. Check the EOL tab — do not order these.</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="success-box">EOL Watch List — All clear.</div>', unsafe_allow_html=True)
            if transfers_enabled and stats["transfers"] > 0:
                st.markdown(f'<div class="transfer-box">{stats["transfers"]} transfers identified — {stats["total_xfer"]} units can be moved internally, {stats["remaining_orders"]} still need external purchase.</div>', unsafe_allow_html=True)
            elif transfers_enabled:
                st.markdown('<div class="info-box">No transfer opportunities found this period.</div>', unsafe_allow_html=True)
            st.markdown("---")
            fname = f"{market_name.strip().replace(' ','_')}_Inventory_Report_{report_date.strftime('%m_%d_%Y')}.xlsx"
            st.download_button(label="Download Excel Report", data=buf, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        except Exception as e:
            st.error(f"Something went wrong: {str(e)}")
 
st.markdown('</div>', unsafe_allow_html=True)
st.markdown('<hr style="border:none;border-top:1px solid #E2E8F0;margin:2rem 0 1rem;"><p style="text-align:center;color:#94A3B8;font-size:0.8rem;">Nextar Telecom Group &middot; Inventory Management &middot; Powered by Nextar Operations</p>', unsafe_allow_html=True)
 
